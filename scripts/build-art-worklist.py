"""
Write docs/missing-sticker-art.xlsx: the sticker cards still to draw, as a worklist.

    python scripts/build-art-worklist.py <cards.json> <out.xlsx>

Called by `npm run ratings:sync`, which hands over the cards as JSON because the dataset
lives in TypeScript and this side cannot read it. Not meant to be run by hand: the whole
point is that the spreadsheet is rewritten every time the ratings move, so it can never
sit there stating a stale count. It replaced a version that was written once and was two
days later wrong in six rows.

Two sheets, as before: the cards as a filterable table, and a summary whose figures are
COUNTA / COUNTIF over that table rather than typed-in numbers, so they follow the rows if
anyone edits them. A reader that shows cached values rather than recalculating will show
those cells blank until the file has been opened in a spreadsheet application once.

Requires openpyxl (pip install openpyxl). ratings:sync reports a missing one and carries
on rather than failing, since a stale worklist breaks nothing.
"""

import json
import os
import sys

try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    sys.exit("build-art-worklist: needs openpyxl.  pip install openpyxl")

if len(sys.argv) != 3:
    sys.exit("build-art-worklist: usage: build-art-worklist.py <cards.json> <out.xlsx>")

SRC, OUT = sys.argv[1], sys.argv[2]

with open(SRC, encoding="utf-8") as fh:
    data = json.load(fh)

cards = data["cards"]
accepted = set(data.get("accepted", []))
fingerprint = data["fingerprint"]

# A spreadsheet is a zip of timestamped XML, so writing an identical one still produces a
# different file and a dirty tree. The fingerprint of what the sheet SAYS lives in it (in
# Summary!B5), so a run with nothing to change leaves the file alone. Recorded on the
# CONTENT only, never the commit, or every commit would rewrite it.
if os.path.exists(OUT):
    try:
        if load_workbook(OUT)["Summary"]["B5"].value == fingerprint:
            print(f"build-art-worklist: {OUT} - unchanged, {len(cards)} cards to draw")
            sys.exit(0)
    except (KeyError, OSError, ValueError):
        pass  # unreadable or an older layout: rewrite it

# The turf-flat pitch-dark, so the sheet is recognisably the same product as the app.
HEAD_FILL = PatternFill("solid", fgColor="1E3A2A")
HEAD_FONT = Font(bold=True, color="FFFFFF")
CENTRE = Alignment(horizontal="center")
WRAP = Alignment(horizontal="left", vertical="top", wrap_text=True)

COLUMNS = [
    ("Player", 24, "left"),
    ("No.", 6, "center"),
    ("Nation", 16, "left"),
    ("World Cup", 12, "center"),
    ("Rating", 9, "center"),
    ("Tier", 14, "left"),
    ("Accepted", 11, "center"),
    ("File to add", 18, "left"),
]

wb = Workbook()
ws = wb.active
ws.title = "Missing art"

for i, (label, width, _align) in enumerate(COLUMNS, start=1):
    cell = ws.cell(row=1, column=i, value=label)
    cell.fill = HEAD_FILL
    cell.font = HEAD_FONT
    cell.alignment = CENTRE if _align == "center" else Alignment(horizontal="left")
    ws.column_dimensions[get_column_letter(i)].width = width

for r, card in enumerate(cards, start=2):
    # "Accepted" is the honest distinction between a gap that is known and shipping as a
    # silhouette (it is on art/awaiting-artwork.txt) and one that is currently failing the
    # build. Without it the sheet cannot tell the two apart, and they need different action.
    values = [
        card["name"],
        card["number"],
        card["nation"],
        card["year"],
        card["rating"],
        card["tier"].capitalize(),
        "yes" if card["id"] in accepted else "NO - failing",
        f"{card['id']}.webp",
    ]
    for i, (value, (_label, _width, align)) in enumerate(zip(values, COLUMNS), start=1):
        cell = ws.cell(row=r, column=i, value=value)
        if align == "center":
            cell.alignment = CENTRE
        if isinstance(value, int):
            cell.number_format = "0"

last = len(cards) + 1
ws.freeze_panes = "A2"
if cards:
    ws.auto_filter.ref = f"A1:{get_column_letter(len(COLUMNS))}{last}"

# --- Summary -----------------------------------------------------------------------------
s = wb.create_sheet("Summary")
s.column_dimensions["A"].width = 34
s.column_dimensions["B"].width = 24
for col in "CDEF":
    s.column_dimensions[col].width = 14

title = s.cell(row=1, column=1, value="Sticker art still to draw")
title.font = Font(bold=True, size=14)

rng = f"'Missing art'!$A$2:$A${last}"
tiers = f"'Missing art'!$F$2:$F${last}"

rows = [
    (3, "Card set last changed", data["generated"]),
    (4, "with the dataset at commit", data["commit"]),
    (5, "Fingerprint", fingerprint),
    (6, "Collectibles in the dataset", data["collectibles"]),
    (7, "With artwork", "=B6-B8"),
    (8, "Missing (rows on the list)", f"=COUNTA({rng})"),
    (9, "Of those, accepted as silhouettes", f"=COUNTIF('Missing art'!$G$2:$G${last},\"yes\")"),
]
for row, label, value in rows:
    s.cell(row=row, column=1, value=label).font = Font(bold=(row in (6, 8)))
    s.cell(row=row, column=2, value=value)

s.cell(row=11, column=1, value="Missing by tier").font = Font(bold=True)
for i, tier in enumerate(["Monumental", "Iconic", "Legendary"]):
    s.cell(row=12 + i, column=1, value=tier)
    s.cell(row=12 + i, column=2, value=f'=COUNTIF({tiers},A{12 + i})')
s.cell(row=15, column=1, value="Total")
s.cell(row=15, column=2, value="=SUM(B12:B14)")

# A card is one player at one tournament, so the same man can owe two. Derived rather than
# written down, because the list changes with every rating pass.
seen: dict[str, list[int]] = {}
for card in cards:
    seen.setdefault(card["name"], []).append(card["year"])
twice = sorted(
    f"{name} ({' and '.join(str(y) for y in sorted(years))})"
    for name, years in seen.items()
    if len(years) > 1
)

notes = [
    "Generated by `npm run ratings:sync`. Do not edit by hand: the next rating change "
    "rewrites this file, and every figure above is derived from the dataset at the commit "
    "named in B4.",
    "To draw a card: save the full-size PNG as art/stickers-src/<id>.png, where <id> is the "
    "file name in the last column minus .webp, then run `npm run ratings:sync`. It builds "
    "the small WebP the site serves and takes the card off the waiting list.",
    "A row saying NO in Accepted is failing the build right now: either draw it or accept "
    "it as a silhouette with `npm run ratings:sync -- --accept-all`.",
]
if twice:
    notes.append(
        "A card is one player at one tournament, so the same man needs a card for each. "
        "On this list: " + ", ".join(twice) + "."
    )

s.cell(row=17, column=1, value="Notes").font = Font(bold=True)
row = 18
for note in notes:
    cell = s.cell(row=row, column=1, value=note)
    cell.alignment = WRAP
    s.merge_cells(start_row=row, start_column=1, end_row=row + 2, end_column=6)
    row += 4

wb.save(OUT)
print(f"build-art-worklist: {OUT} - {len(cards)} cards to draw, {len(accepted)} accepted")
